# =================================================================================
#   ФИНАЛЬНАЯ ВЕРСИЯ БОТА (V42 - ИСПРАВЛЕНИЕ СХЕМЫ БД)
# =================================================================================

# --- 1. ИМПОРТЫ ---
import os
import logging
import asyncio
from contextlib import asynccontextmanager
from datetime import datetime
import zipfile
import io
from typing import List, Dict, Any, Optional, Tuple, Set
import psycopg2
import uuid
import re

import uvicorn
from fastapi import FastAPI
from telegram import Update, ReplyKeyboardMarkup, Message, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
    ConversationHandler,
    CallbackQueryHandler,
)
from cryptography import x509
from cryptography.hazmat.backends import default_backend
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

# --- 2. НАСТРОЙКА И КОНСТАНТЫ ---
TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN")
DATABASE_URL = os.environ.get("DATABASE_URL")

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

ALLOWED_USER_IDS: Set[int] = {96238783}
user_filter = filters.User(user_id=ALLOWED_USER_IDS)

MAX_FILE_SIZE = 20 * 1024 * 1024
EXPIRATION_THRESHOLD_DAYS = 30
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFDDAA", end_color="FFDDAA", fill_type="solid")
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
EXCEL_HEADERS: Tuple[str, ...] = ("ФИО", "Учреждение", "Серийный номер", "Действителен с", "Действителен до", "Осталось дней")
ALLOWED_EXTENSIONS: Tuple[str, ...] = ('.cer', '.crt', '.pem', '.der')
YOUTUBE_URL_PATTERN = r'(https?://)?(www\.)?(youtube|youtu|youtube-nocookie)\.(com|be)/(watch\?v=|embed/|v/|.+\?v=)?([^&=%\?]{11})'
CHOOSING_ACTION, TYPING_DAYS = range(2)

# --- 3. ИНИЦИАЛИЗАЦИЯ БОТА И ВЕБ-СЕРВЕРА ---

# Создаем объект application глобально, чтобы он был доступен в lifespan
if not TELEGRAM_BOT_TOKEN:
    logger.error("Токен Telegram бота не найден! Завершение работы.")
    exit()
application = Application.builder().token(TELEGRAM_BOT_TOKEN).build()

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Управляет жизненным циклом бота вместе с веб-сервером."""
    logger.info("Lifespan: Запуск бота...")
    if not DATABASE_URL:
        logger.warning("DATABASE_URL не найден. Работа с БД будет невозможна.")
    else:
        init_database()
        
    await application.initialize()
    await application.start()
    await application.updater.start_polling()
    yield
    logger.info("Lifespan: Остановка бота...")
    await application.updater.stop()
    await application.stop()
    await application.shutdown()

# Создаем FastAPI приложение с lifespan
app = FastAPI(lifespan=lifespan, docs_url=None, redoc_url=None)

@app.get("/")
async def root():
    return {"status": "bot is running"}


# --- 4. РАБОТА С БАЗОЙ ДАННЫХ ---
def get_db_connection():
    try:
        conn = psycopg2.connect(DATABASE_URL)
        return conn
    except Exception as e:
        logger.error(f"Не удалось подключиться к базе данных: {e}")
        return None

def init_database():
    conn = get_db_connection()
    if not conn: return
    try:
        with conn.cursor() as cursor:
            cursor.execute('CREATE TABLE IF NOT EXISTS user_settings (user_id BIGINT PRIMARY KEY, threshold INTEGER NOT NULL)')
            
            # <<< ИСПРАВЛЕНИЕ: Добавлена колонка local_filepath >>>
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS download_tasks (
                    task_id UUID PRIMARY KEY,
                    user_id BIGINT NOT NULL,
                    youtube_url TEXT NOT NULL,
                    status VARCHAR(20) DEFAULT 'new',
                    local_filepath TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
        conn.commit()
        logger.info("База данных PostgreSQL успешно инициализирована.")
    except Exception as e:
        logger.error(f"Ошибка при инициализации таблиц: {e}")
    finally:
        if conn: conn.close()

def save_user_threshold(user_id: int, threshold: int):
    conn = get_db_connection()
    if not conn: return
    try:
        with conn.cursor() as cursor:
            cursor.execute("INSERT INTO user_settings (user_id, threshold) VALUES (%s, %s) ON CONFLICT (user_id) DO UPDATE SET threshold = EXCLUDED.threshold;",(user_id, threshold))
        conn.commit()
    finally:
        if conn: conn.close()

def load_user_threshold(user_id: int) -> Optional[int]:
    conn = get_db_connection()
    if not conn: return None
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT threshold FROM user_settings WHERE user_id = %s", (user_id,))
            result = cursor.fetchone()
        return result[0] if result else None
    finally:
        if conn: conn.close()

async def get_user_threshold(user_id: int, context: ContextTypes.DEFAULT_TYPE) -> int:
    if 'threshold' in context.user_data:
        return context.user_data['threshold']
    threshold_from_db = load_user_threshold(user_id)
    if threshold_from_db is not None:
        context.user_data['threshold'] = threshold_from_db
        return threshold_from_db
    return EXPIRATION_THRESHOLD_DAYS

def create_download_task(user_id: int, youtube_url: str) -> Optional[str]:
    conn = get_db_connection()
    if not conn: return None
    task_id = uuid.uuid4()
    try:
        with conn.cursor() as cursor:
            cursor.execute("INSERT INTO download_tasks (task_id, user_id, youtube_url) VALUES (%s, %s, %s)", (str(task_id), user_id, youtube_url))
        conn.commit()
        logger.info(f"Создано задание {task_id} для пользователя {user_id}")
        return str(task_id)
    except psycopg2.Error as e:
        logger.error(f"Ошибка при создании задания: {e}", exc_info=True)
        return None
    finally:
        if conn: conn.close()


# --- 5. ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ И ОБРАБОТЧИКИ ---
def create_excel_report(cert_data_list: List[Dict[str, Any]], user_threshold: int) -> io.BytesIO:
    wb = Workbook(); ws = wb.active; ws.title = "Отчет по сертификатам"
    ws.append(list(EXCEL_HEADERS)); sorted_cert_data = sorted(cert_data_list, key=lambda x: x["Действителен до"])
    for cert_data in sorted_cert_data:
        row = [cert_data["ФИО"], cert_data["Учреждение"], cert_data["Серийный номер"], cert_data["Действителен с"].strftime("%d.%m.%Y"), cert_data["Действителен до"].strftime("%d.%m.%Y"), cert_data["Осталось дней"]]
        ws.append(row); last_row = ws.max_row; days_left = cert_data["Осталось дней"]
        fill_color = None
        if days_left < 0: fill_color = RED_FILL
        elif 0 <= days_left <= user_threshold: fill_color = ORANGE_FILL
        else: fill_color = GREEN_FILL
        if fill_color:
            for cell in ws[last_row]: cell.fill = fill_color
    for column in ws.columns:
        max_length = 0; column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2); ws.column_dimensions[column_letter].width = adjusted_width
    excel_buffer = io.BytesIO(); wb.save(excel_buffer); excel_buffer.seek(0)
    return excel_buffer

def generate_summary_message(cert_data_list: List[Dict[str, Any]], user_threshold: int) -> str:
    expiring_soon_certs = []
    for cert_data in cert_data_list:
        days_left = cert_data["Осталось дней"]
        if 0 <= days_left <= user_threshold:
            expiring_soon_certs.append(f"👤 {cert_data['ФИО']} — {cert_data['Действителен до'].strftime('%d.%m.%Y')} (осталось {days_left} дн.)")
    if expiring_soon_certs:
        message_parts = [f"⚠️ Скоро истекают ({user_threshold} дней):", *expiring_soon_certs]
        return "\n".join(message_parts)
    else: return "✅ Сертификатов, истекающих в ближайшее время, не найдено."

def get_certificate_info(cert_bytes: bytes) -> Optional[Dict[str, Any]]:
    try:
        try: cert = x509.load_pem_x509_certificate(cert_bytes, default_backend())
        except ValueError: cert = x509.load_der_x509_certificate(cert_bytes, default_backend())
        try: subject_common_name = cert.subject.get_attributes_for_oid(x509.OID_COMMON_NAME)[0].value
        except IndexError: subject_common_name = "Неизвестно"
        try: organization_name = cert.subject.get_attributes_for_oid(x509.OID_ORGANIZATION_NAME)[0].value
        except IndexError: organization_name = "Неизвестно"
        serial_number = f"{cert.serial_number:X}"; valid_from = cert.not_valid_before.date(); valid_until = cert.not_valid_after.date()
        days_left = (valid_until - datetime.now().date()).days
        return {"ФИО": subject_common_name, "Учреждение": organization_name, "Серийный номер": serial_number, "Действителен с": valid_from, "Действителен до": valid_until, "Осталось дней": days_left}
    except Exception as e:
        logger.error(f"Ошибка при парсинге сертификата: {e}"); return None

def _process_file_content(file_bytes: bytes, file_name: str) -> List[Dict[str, Any]]:
    all_certs_data = []
    if file_name.lower().endswith(".zip"):
        try:
            with zipfile.ZipFile(io.BytesIO(file_bytes), 'r') as z:
                for member in z.namelist():
                    if member.lower().endswith(ALLOWED_EXTENSIONS):
                        with z.open(member) as cert_file:
                            cert_info = get_certificate_info(cert_file.read())
                            if cert_info: all_certs_data.append(cert_info)
        except zipfile.BadZipFile:
            logger.error(f"Получен поврежденный ZIP-файл: {file_name}", exc_info=True); return []
    elif file_name.lower().endswith(ALLOWED_EXTENSIONS):
        cert_info = get_certificate_info(file_bytes)
        if cert_info: all_certs_data.append(cert_info)
    return all_certs_data

async def get_my_id(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    await update.message.reply_text(f"Ваш User ID: `{user_id}`", parse_mode='Markdown')

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    keyboard = [["📜 Сертификат", "📄 Заявка АКЦ"], ["⚙️ Настройки", "❓ Помощь"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    start_message = (f"Привет, {user.mention_html()}! 👋\n\nЯ бот для анализа цифровых сертификатов и создания задач на скачивание видео.")
    await update.message.reply_html(start_message, reply_markup=reply_markup)

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text("Чтобы получить отчет по сертификатам, отправьте файлы. Чтобы скачать видео, отправьте ссылку на YouTube.")

async def request_certificate_files(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(f"Пожалуйста, отправьте мне файл(ы) сертификатов ({', '.join(ALLOWED_EXTENSIONS)}) или ZIP-архив.")

async def acc_finance_placeholder(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_html("📈 **Функция 'Заявка АКЦ-Финансы' в разработке.**")

async def handle_simple_buttons(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    button_text = update.message.text
    if button_text == "❓ Помощь":
        await help_command(update, context)
    elif button_text == "📜 Сертификат":
        await request_certificate_files(update, context)
    elif button_text == "📄 Заявка АКЦ":
        await acc_finance_placeholder(update, context)

async def handle_youtube_link(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    url = update.message.text
    task_id = create_download_task(update.effective_user.id, url)
    if task_id:
        await update.message.reply_text(f"✅ Задание на скачивание создано.\n\nID: `{task_id}`\n\nЗапустите `worker.py` на вашем компьютере.", parse_mode='Markdown')
    else:
        await update.message.reply_text("❌ Не удалось создать задание на скачивание.")

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    document = update.message.document
    if document.file_size > MAX_FILE_SIZE:
        await update.message.reply_text(f"❌ Файл слишком большой. Максимум: {MAX_FILE_SIZE / 1024 / 1024:.0f} МБ."); return
    user_id = update.effective_user.id; user_threshold = await get_user_threshold(user_id, context)
    file_name = document.file_name; logger.info(f"Получен файл: {file_name} от {user_id}")
    await update.message.reply_text("Анализирую...")
    try:
        file_object = await context.bot.get_file(document.file_id); file_buffer = io.BytesIO()
        await file_object.download_to_memory(file_buffer); file_buffer.seek(0)
        all_certs_data = _process_file_content(file_buffer.read(), file_name)
        if not all_certs_data:
            await update.message.reply_text("Не удалось найти/проанализировать сертификаты."); return
        excel_buffer = create_excel_report(all_certs_data, user_threshold); summary_message = generate_summary_message(all_certs_data, user_threshold)
        await update.message.reply_text(summary_message); await update.message.reply_document(document=excel_buffer, filename="Сертификаты_отчет.xlsx")
    except Exception as e:
        logger.error(f"Ошибка при обработке документа: {e}", exc_info=True); await update.message.reply_text(f"Произошла непредвиденная ошибка.")

async def handle_wrong_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(f"❌ Неверный формат файла. Нужны: {', '.join(ALLOWED_EXTENSIONS)}, .zip")

async def settings_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id; current_threshold = await get_user_threshold(user_id, context)
    keyboard = [[InlineKeyboardButton("Изменить порог", callback_data='change_threshold')], [InlineKeyboardButton("Назад", callback_data='back_to_main')]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(f"⚙️ **Настройки**\nТекущий порог: **{current_threshold}** дней.", reply_markup=reply_markup, parse_mode='Markdown')
    return CHOOSING_ACTION

async def prompt_for_days(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query; await query.answer()
    await query.edit_message_text(text="Отправьте новое число дней (например, 60).")
    return TYPING_DAYS

async def set_days(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    try:
        new_threshold = int(update.message.text)
        if new_threshold <= 0:
            await update.message.reply_text("❌ Введите положительное число."); return TYPING_DAYS
        context.user_data['threshold'] = new_threshold; save_user_threshold(user_id, new_threshold)
        await update.message.reply_html(f"✅ Порог изменен и сохранен: <b>{new_threshold}</b> дней.")
    except (ValueError):
        await update.message.reply_text("❌ Это не число. Отправьте, например: 60"); return TYPING_DAYS
    return ConversationHandler.END

async def end_conversation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query; await query.answer()
    await query.edit_message_text(text="Настройки закрыты.")
    return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text('Действие отменено.')
    return ConversationHandler.END


# --- 6. РЕГИСТРАЦИЯ ОБРАБОТЧИКОВ ---
settings_conv_handler = ConversationHandler(
    entry_points=[MessageHandler(filters.Regex('^⚙️ Настройки$') & user_filter, settings_entry)],
    states={
        CHOOSING_ACTION: [CallbackQueryHandler(prompt_for_days, pattern='^change_threshold$'), CallbackQueryHandler(end_conversation, pattern='^back_to_main$')],
        TYPING_DAYS: [MessageHandler(filters.TEXT & ~filters.COMMAND, set_days)],
    },
    fallbacks=[CommandHandler('start', start), MessageHandler(filters.Regex('^(📜 Сертификат|📄 Заявка АКЦ|❓ Помощь)$'), cancel)],
)

application.add_handler(settings_conv_handler)
application.add_handler(CommandHandler("my_id", get_my_id))
application.add_handler(CommandHandler("start", start, filters=user_filter))
application.add_handler(MessageHandler(filters.Regex(YOUTUBE_URL_PATTERN) & user_filter, handle_youtube_link))
simple_buttons_text = "^(📜 Сертификат|📄 Заявка АКЦ|❓ Помощь)$"
application.add_handler(MessageHandler(filters.Regex(simple_buttons_text) & user_filter, handle_simple_buttons))
allowed_extensions_filter = (
    filters.Document.FileExtension("zip") | filters.Document.FileExtension("cer") |
    filters.Document.FileExtension("crt") | filters.Document.FileExtension("pem") |
    filters.Document.FileExtension("der")
)
application.add_handler(MessageHandler(allowed_extensions_filter & ~filters.COMMAND & user_filter, handle_document))
application.add_handler(MessageHandler(filters.Document.ALL & ~filters.COMMAND & user_filter, handle_wrong_document))


# --- 7. ТОЧКА ВХОДА ---
if __name__ == "__main__":
    # Эта часть теперь нужна только для тестов на вашем компьютере.
    # Render будет использовать команду uvicorn и не будет выполнять этот блок.
    logger.info("Запуск в локальном режиме...")
    uvicorn.run(app, host="0.0.0.0", port=8000)
